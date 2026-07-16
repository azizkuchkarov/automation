import sys
from openpyxl import load_workbook
path = sys.argv[1]
ws = load_workbook(path).active
lines = []
for r in range(2, 11):
    rd = ws.row_dimensions[r]
    lines.append(f"row {r}: height={rd.height}, hidden={rd.hidden}")
for addr in ["D4", "D7", "E2", "F8"]:
    c = ws[addr]
    a = c.alignment
    lines.append(f"{addr}: v={c.alignment.vertical}, h={c.alignment.horizontal}, wrap={c.alignment.wrap_text}, len={len(str(c.value or ''))}")
open(sys.argv[2], "w", encoding="utf-8").write("\n".join(lines))
