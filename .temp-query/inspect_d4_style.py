import sys
from openpyxl import load_workbook
from openpyxl.styles import Color

path = sys.argv[1]
ws = load_workbook(path).active
c = ws["D4"]
fill = c.fill
font = c.font
lines = [
    f"D4 value: {c.value!r}",
    f"font color: {font.color.rgb if font.color else None}",
    f"font bold: {font.bold}",
    f"fill type: {fill.patternType}",
    f"fill fg: {fill.fgColor.rgb if fill.fgColor else None}",
    f"fill bg: {fill.bgColor.rgb if fill.bgColor else None}",
    f"row4 hidden: {ws.row_dimensions[4].hidden}",
    f"row4 height: {ws.row_dimensions[4].height}",
    f"alignment wrap: {c.alignment.wrap_text}",
    f"alignment shrink: {c.alignment.shrink_to_fit}",
]
open(sys.argv[2], "w", encoding="utf-8").write("\n".join(lines))
