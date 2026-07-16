from openpyxl import load_workbook

path = r"C:\Users\a.kuchkarov\Desktop\automation\ATG.Platform.Infrastructure\Hr\Templates\BusinessTripCertificate.xlsx"
wb = load_workbook(path)
ws = wb.active
ws["D4"] = "должность / position"
ws.row_dimensions[4].hidden = False
ws.row_dimensions[4].height = None
wb.save(path)
print("template updated")
